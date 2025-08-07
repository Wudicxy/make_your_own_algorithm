# views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, F
from django.contrib import messages
from .models import Question, Category, Difficulty, UserFavorite
# views.py (添加到原有views.py中)
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.http import HttpResponse
from .forms import ExcelImportForm, JSONImportForm, ManualBatchForm, QuickAddForm
from .services import QuestionImportService
from .models import ImportRecord
import openpyxl
from django.shortcuts import get_object_or_404

def import_record_detail(request, pk):
    """导入记录详情"""
    import_record = get_object_or_404(ImportRecord, pk=pk)
    return render(request, 'baguwen/import_record_detail.html', {'import_record': import_record})

def import_questions(request):
    """题目导入主页"""
    import_records = ImportRecord.objects.all()[:10]  # 最近10条导入记录

    context = {
        'import_records': import_records,
        'excel_form': ExcelImportForm(),
        'json_form': JSONImportForm(),
        'manual_form': ManualBatchForm(),
        'quick_form': QuickAddForm(),
    }

    return render(request, 'baguwen/import_questions.html', context)


@login_required
def import_excel(request):
    """Excel导入"""
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                service = QuestionImportService(request.user)
                import_record = service.import_from_excel(request.FILES['excel_file'])

                messages.success(
                    request,
                    f'导入完成！成功：{import_record.success_count}，失败：{import_record.error_count}'
                )

                if import_record.error_count > 0:
                    messages.warning(request, f'部分导入失败，请查看导入记录了解详情')

            except Exception as e:
                messages.error(request, f'导入失败：{str(e)}')

    return redirect('baguwen:import_questions')


@login_required
def import_json(request):
    """JSON导入"""
    if request.method == 'POST':
        form = JSONImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                service = QuestionImportService(request.user)
                import_record = service.import_from_json(request.FILES['json_file'])

                messages.success(
                    request,
                    f'导入完成！成功：{import_record.success_count}，失败：{import_record.error_count}'
                )

            except Exception as e:
                messages.error(request, f'导入失败：{str(e)}')

    return redirect('baguwen:import_questions')


@login_required
def import_manual(request):
    """手动批量导入"""
    if request.method == 'POST':
        form = ManualBatchForm(request.POST)
        if form.is_valid():
            try:
                service = QuestionImportService(request.user)
                import_record = service.import_from_manual_data(form.cleaned_data['questions_data'])

                messages.success(
                    request,
                    f'导入完成！成功：{import_record.success_count}，失败：{import_record.error_count}'
                )

            except Exception as e:
                messages.error(request, f'导入失败：{str(e)}')

    return redirect('baguwen:import_questions')


@login_required
def quick_add(request):
    """快速添加单个题目"""
    if request.method == 'POST':
        form = QuickAddForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '题目添加成功！')
            return redirect('baguwen:quick_add')
    else:
        form = QuickAddForm()

    return render(request, 'baguwen/quick_add.html', {'form': form})


def download_template(request):
    """下载Excel导入模板"""
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "题目导入模板"

    # 设置表头
    headers = ['title', 'content', 'answer', 'key_points', 'category', 'difficulty', 'tags']
    header_names = ['题目标题', '题目内容', '参考答案', '关键要点', '分类', '难度', '标签']

    for i, (header, name) in enumerate(zip(headers, header_names), 1):
        ws.cell(row=1, column=i, value=f'{name}({header})')

    # 添加示例数据
    example_data = [
        ['什么是TCP三次握手？', 'TCP连接建立过程的详细说明', 'TCP三次握手是...',
         'SYN, ACK, 连接状态', '网络协议', '中级', 'TCP,网络,协议'],
        ['解释HTTP和HTTPS的区别', '请详细说明两者的不同点', 'HTTP是明文传输...',
         '加密, SSL/TLS, 端口', 'Web技术', '初级', 'HTTP,HTTPS,安全']
    ]

    for row_num, row_data in enumerate(example_data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # 设置响应
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="baguwen_import_template.xlsx"'
    wb.save(response)

    return response


def import_record_detail(request, pk):
    """导入记录详情"""
    import_record = get_object_or_404(ImportRecord, pk=pk)
    return render(request, 'baguwen/import_record_detail.html', {'import_record': import_record})


# def question_list(request):
#     """题目列表页"""
#     questions = Question.objects.filter(is_active=True)
#
#     # 搜索功能
#     search_query = request.GET.get('search', '')
#     if search_query:
#         questions = questions.filter(
#             Q(title__icontains=search_query) |
#             Q(content__icontains=search_query) |
#             Q(tags__icontains=search_query)
#         )
#
#     # 分类筛选
#     category_id = request.GET.get('category')
#     if category_id:
#         questions = questions.filter(category_id=category_id)
#
#     # 难度筛选
#     difficulty_id = request.GET.get('difficulty')
#     if difficulty_id:
#         questions = questions.filter(difficulty_id=difficulty_id)
#
#     # 分页
#     paginator = Paginator(questions, 10)
#     page_number = request.GET.get('page')
#     page_obj = paginator.get_page(page_number)
#
#     # 获取所有分类和难度用于筛选
#     categories = Category.objects.all()
#     difficulties = Difficulty.objects.all().order_by('level')
#
#     context = {
#         'page_obj': page_obj,
#         'categories': categories,
#         'difficulties': difficulties,
#         'search_query': search_query,
#         'selected_category': category_id,
#         'selected_difficulty': difficulty_id,
#     }
#
#     return render(request, 'baguwen/questions_bagu_list.html', context)
# views.py
def question_list(request):
    """题目列表页"""
    questions = Question.objects.filter(is_active=True)

    # 搜索功能
    search_query = request.GET.get('search', '')
    if search_query:
        questions = questions.filter(
            Q(title__icontains=search_query) |
            Q(content__icontains=search_query) |
            Q(tags__icontains=search_query)
        )

    # 分类筛选
    category_id = request.GET.get('category')
    if category_id:
        questions = questions.filter(category_id=category_id)

    # 难度筛选
    difficulty_id = request.GET.get('difficulty')
    if difficulty_id:
        questions = questions.filter(difficulty_id=difficulty_id)

    # 分页
    paginator = Paginator(questions, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 获取所有分类和难度用于筛选
    categories = Category.objects.all()
    difficulties = Difficulty.objects.all().order_by('level')

    # 构建保持筛选条件的查询字符串
    query_params = request.GET.copy()
    if 'page' in query_params:
        del query_params['page']
    query_string = query_params.urlencode()

    context = {
        'page_obj': page_obj,
        'categories': categories,
        'difficulties': difficulties,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_difficulty': difficulty_id,
        'query_string': query_string,  # 新增
    }

    return render(request, 'baguwen/questions_bagu_list.html', context)


def question_detail(request, pk):
    """题目详情页"""
    question = get_object_or_404(Question, pk=pk, is_active=True)

    # 增加浏览次数
    Question.objects.filter(pk=pk).update(view_count=F('view_count') + 1)

    # 检查是否已收藏
    is_favorited = False
    if request.user.is_authenticated:
        is_favorited = UserFavorite.objects.filter(
            user=request.user, question=question
        ).exists()

    # 获取相关题目
    related_questions = Question.objects.filter(
        category=question.category,
        is_active=True
    ).exclude(pk=pk)[:5]

    context = {
        'question': question,
        'is_favorited': is_favorited,
        'related_questions': related_questions,
    }

    return render(request, 'baguwen/question_detail.html', context)


def category_list(request):
    """分类列表页"""
    categories = Category.objects.all()

    # 为每个分类添加题目数量
    for category in categories:
        category.question_count = Question.objects.filter(
            category=category, is_active=True
        ).count()

    context = {'categories': categories}
    return render(request, 'baguwen/category_list.html', context)


@login_required
def toggle_favorite(request, pk):
    """切换收藏状态"""
    if request.method == 'POST':
        question = get_object_or_404(Question, pk=pk)
        favorite, created = UserFavorite.objects.get_or_create(
            user=request.user, question=question
        )

        if not created:
            favorite.delete()
            is_favorited = False
            message = '已取消收藏'
        else:
            is_favorited = True
            message = '收藏成功'

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'is_favorited': is_favorited,
                'message': message
            })
        else:
            messages.success(request, message)
            return redirect('baguwen:question_detail', pk=pk)

    return redirect('baguwen:question_list')


@login_required
def my_favorites(request):
    """我的收藏页"""
    favorites = UserFavorite.objects.filter(user=request.user).select_related('question')

    # 分页
    paginator = Paginator(favorites, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {'page_obj': page_obj}
    return render(request, 'baguwen/my_favorites.html', context)


def random_question(request):
    """随机题目"""
    question = Question.objects.filter(is_active=True).order_by('?').first()
    if question:
        return redirect('baguwen:question_detail', pk=question.pk)
    else:
        messages.warning(request, '暂无可用题目')
        return redirect('baguwen:question_list')


from django.shortcuts import render

# Create your views here.
